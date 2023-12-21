# %% [markdown]
# #### 爬取专题
# - 同城
# - 明星
# - 搞笑
# - 游戏
# - 美妆
# - 美食
# - 摄影
# - 数码
# - 萌宠
# - 星座
# - 体育
# - 旅游
# - 教育
# - 国际
# - 财经
# - 婚恋
# - 时尚
# - 科技
# - 校园
# - 动漫

# %% [markdown]
# #### 爬取数量
# - 爬取1200位用户
# - 每位用户爬取1000条数据，共360000条数据

# %% [markdown]
# #### 隐藏身份
# 
# - `Android`
# 
#     - Mozilla/5.0 (Linux; Android 4.1.1; Nexus 7 Build/JRO03D) AppleWebKit/535.19 (KHTML, like Gecko) Chrome/18.0.1025.166 Safari/535.19
#     - Mozilla/5.0 (Linux; U; Android 4.0.4; en-gb; GT-I9300 Build/IMM76D) AppleWebKit/534.30 (KHTML, like Gecko) Version/4.0 Mobile Safari/534.30
#     - Mozilla/5.0 (Linux; U; Android 2.2; en-gb; GT-P1000 Build/FROYO) AppleWebKit/533.1 (KHTML, like Gecko) Version/4.0 Mobile Safari/533.1
# 
# - `Firefox`
# 
#     - Mozilla/5.0 (Windows NT 6.2; WOW64; rv:21.0) Gecko/20100101 Firefox/21.0
#     - Mozilla/5.0 (Android; Mobile; rv:14.0) Gecko/14.0 Firefox/14.0
# 
# - `Google Chrome`
# 
#     - Mozilla/5.0 (Windows NT 6.2; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/27.0.1453.94 Safari/537.36
#     - Mozilla/5.0 (Linux; Android 4.0.4; Galaxy Nexus Build/IMM76B) AppleWebKit/535.19 (KHTML, like Gecko) Chrome/18.0.1025.133 Mobile Safari/535.19
# 
# - `iOS`
# 
#     - Mozilla/5.0 (iPad; CPU OS 5_0 like Mac OS X) AppleWebKit/534.46 (KHTML, like Gecko) Version/5.1 Mobile/9A334 Safari/7534.48.3
#     - Mozilla/5.0 (iPod; U; CPU like Mac OS X; en) AppleWebKit/420.1 (KHTML, like Gecko) Version/3.0 Mobile/3A101a Safari/419.3

# %% [markdown]
# #### Firefox设置IP代理
# 
# ```python
#     from selenium import webdriver
# 
#     profile = webdriver.FirefoxProfile()
# 
#     # 激活手动代理配置
#     profile.set_preference("network.proxy.type", 1)
#     profile.set_preference("network.proxy.http", HOST)
#     profile.set_preference("network.proxy.http_port", PORT)
# 
#     # 所有协议共用一种 ip 及端口，如果单独配置，不必设置该项，因为其默认为 False
#     profile.set_preference("network.proxy.share_proxy_settings", True)
# 
#     # 默认本地地址（localhost）不使用代理，如果有些域名在访问时不想使用代理可以使用类似下面的参数设置
#     # profile.set_preference("network.proxy.no_proxies_on", "localhost")
# 
#     profile.set_preference("general.useragent.override", "Mozilla/5.0 ...")
# 
#     driver = webdriver.Firefox(firefox_profile=profile)
# 
#     driver.get("https://weibo.com")
# ```

# %% [markdown]
# #### 初始化环境

# %% [markdown]
# - 添加库

# %%
import traceback
import openpyxl
import requests
import random
import shutil
import json
import copy
import os


from tqdm import tqdm
from time import sleep
from scrapy import Selector
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# %% [markdown]
# - 定义文件属性

# %%
user_fields     =   ['uid',         '专题',             '头像',         '背景', 
                     '用户名',      '性别',             '会员等级',     '粉丝数', 
                     '关注数',      '全部微博',         '自我介绍',     '地域', 
                     'IP',          '加入微博时间',     '信用',         '出生日期/星座', 
                     'v认证',       '身份认证',         '表现',         '内容/IP机构认证',
                     '校园',        '合作机构']


follow_fields   =   ['follower_id', 'following_id',     '头像',         '背景', 
                     '用户名',      '性别',             '会员等级',     '粉丝数',   
                     '关注数',      '全部微博',         '自我介绍',     '地域', 
                     'IP',          '加入微博时间',     '信用',         '出生日期/星座', 
                     'v认证',       '身份认证',         '表现',         '内容/IP机构认证', 
                     '校园',        '合作机构']


weibo_fields    =   ['uid',             '日期',         '来源/设备',    '文本',
                     '图片',            '转发数',       '评论数',       '点赞数',
                     '是否转发',        '是否快转',     '是否点赞',     '原作者uid',    
                     '原作者用户名',    '原发布日期',   '原设备',       '原文本',
                     '原图片',          '原转发数',     '原评论数',     '原点赞数']

# %% [markdown]
# - 构造文件结构

# %%
spider_dir      = 'spider'
graphs_dir      = 'spider/graphs'
user_profiles   = 'spider/user_profiles.xlsx'
follow_profiles = 'spider/follow_profiles.xlsx'
weibo_profiles  = 'spider/weibo.xlsx'
cookies_txt     = 'spider/cookies.txt'

# %%
os.makedirs(spider_dir, mode=777)           if not os.path.exists(spider_dir)       else None
os.makedirs(graphs_dir, mode=777)           if not os.path.exists(graphs_dir)       else None
openpyxl.Workbook().save(user_profiles)     if not os.path.exists(user_profiles)    else None
openpyxl.Workbook().save(follow_profiles)   if not os.path.exists(follow_profiles)  else None
openpyxl.Workbook().save(weibo_profiles)    if not os.path.exists(weibo_profiles)   else None
open(cookies_txt, 'w+').close()             if not os.path.exists(cookies_txt)      else None

# %% [markdown]
# - Excel追加内容

# %%
def add_to_excel(path, lst):
    target_book = openpyxl.load_workbook(path)
    
    for line in lst:
        target_book.active.append(line)

    target_book.save(path)
    target_book.close()

# %% [markdown]
# - 更新cookies

# %%
def update_cookies(path=cookies_txt):
    driver = webdriver.Firefox()
    driver.get('https://weibo.com/login.php')
    while driver.current_url.find('https://weibo.com/login.php') == -1: sleep(3)
    while driver.current_url.find('https://weibo.com/login.php') != -1: sleep(3)
    sleep(3)
    with open(path, 'w+', encoding='utf-8') as f: f.write(json.dumps(driver.get_cookies()))
    driver.close()
    driver.quit()

# %% [markdown]
# - 配置cookies

# %%
# 设置 cookies
def set_cookies(driver, path=cookies_txt):
    with open(path, 'r+', encoding='utf-8') as f: cookies = eval(f.read().replace('false', 'False').replace('true', 'True'))
    for cookie in cookies:
        conf = dict({   'domain'    :cookie['domain'], 
                        'name'      :cookie['name'], 
                        'value'     :cookie['value']}, 
                  **{   'expires'   :'', 
                        'path'      :'/', 
                        'httpOnly'  :False, 
                        'HostOnly'  :False, 
                        'Secure'    :False})
        driver.add_cookie(conf)

# %% [markdown]
# #### 定义函数、实现爬虫

# %% [markdown]
# ```python
# 'get_user_fields(driver, href)'
# ```
# 
# 参数|说明
# ----|----
# driver | 驱动对象
# href | 用户的主页链接
# 说明 | 只适用于<a>https://weibo.com</a>

# %%
def get_user_fields(driver, href):

    """声明变量"""
    avatar_url,     background_url,     username,       gender,                 \
    level,          followers,          followings,     weibo_count,            \
    introduction,   address,            ip,             join_date,              \
    credit,         birth,              v_auth,         id_auth,                \
    performance,    content_auth,       campus,         cooperative_org   =   [None]*20
    
    uid  = href.replace('https://weibo.com/u/', '')

    # 进入用户主页
    while driver.current_url != href:
        try: driver.get(href); break
        except: traceback

        try: driver.get(random.choice(['https://weibo.com', 'https://weibo.com/at/weibo']))
        except: traceback
        sleep(random.randint(20, 60))


    while True:
        try: driver.find_element(By.XPATH, "//i[@class='woo-font woo-font--angleDown']").click(); break
        except: traceback

    response        = Selector(text=driver.page_source)

    # 获取头像
    avatar_url      = response.xpath("//div[@class='woo-panel-main woo-panel-top woo-panel-right woo-panel-bottom woo-panel-left Card_wrap_2ibWe Card_bottomGap_2Xjqi']//div[@class='woo-avatar-main woo-avatar-hover ProfileHeader_avatar2_1gEyo']/img/@src").extract_first()
    
    # 获取背景
    background_url  = response.xpath("//div[@class='woo-panel-main woo-panel-top woo-panel-right woo-panel-bottom woo-panel-left Card_wrap_2ibWe Card_bottomGap_2Xjqi']//div[@class='woo-picture-main ProfileHeader_pic_2Coeq']/img/@src").extract_first()

    # 获取用户名
    username        = response.xpath("//div[@class='woo-panel-main woo-panel-top woo-panel-right woo-panel-bottom woo-panel-left Card_wrap_2ibWe Card_bottomGap_2Xjqi']//div[@class='ProfileHeader_name_1KbBs']/text()").extract_first()

    # 获取性别
    gender          = 'female' if response.xpath("//div[@class='woo-panel-main woo-panel-top woo-panel-right woo-panel-bottom woo-panel-left Card_wrap_2ibWe Card_bottomGap_2Xjqi']//svg[@class='woo-icon-main woo-icon--female']").extract_first() else 'male'

    # 获取会员等级
    level           = response.xpath("//div[@class='woo-panel-main woo-panel-top woo-panel-right woo-panel-bottom woo-panel-left Card_wrap_2ibWe Card_bottomGap_2Xjqi']//span[@class='woo-icon-wrap IconVip_icon_2tjdp']/@aria-label").extract_first()

    # 获取粉丝数
    followers       = response.xpath(f"//div[@class='woo-panel-main woo-panel-top woo-panel-right woo-panel-bottom woo-panel-left Card_wrap_2ibWe Card_bottomGap_2Xjqi']//a[@class='ALink_none_1w6rm ProfileHeader_alink_tjHJR ProfileHeader_pointer_2yKGQ' and @href='/u/page/follow/{uid}?relate=fans']/span/span/text()").extract_first()
    
    # 获取关注数
    followings      = response.xpath(f"//div[@class='woo-panel-main woo-panel-top woo-panel-right woo-panel-bottom woo-panel-left Card_wrap_2ibWe Card_bottomGap_2Xjqi']//a[@class='ALink_none_1w6rm ProfileHeader_alink_tjHJR ProfileHeader_pointer_2yKGQ' and @href='/u/page/follow/{uid}?relate=']/span/span/text()").extract_first()

    # 获取全部微博数量
    weibo_count     = response.xpath("//div[@class='wbpro-screen-v2 woo-box-flex woo-box-alignCenter woo-box-justifyBetween']/div/text()").extract_first()

    # 获取用户自我介绍
    introduction    = response.xpath("//div[@class='woo-panel-main woo-panel-top woo-panel-right woo-panel-bottom woo-panel-left Card_wrap_2ibWe Card_bottomGap_2Xjqi']//i[@class='woo-font woo-font--proBintro']/../../div[@class='woo-box-item-flex ProfileHeader_con3_Bg19p']/text()").extract_first()
    
    # 获取用户地域
    address         = response.xpath("//div[@class='woo-panel-main woo-panel-top woo-panel-right woo-panel-bottom woo-panel-left Card_wrap_2ibWe Card_bottomGap_2Xjqi']//i[@class='woo-font woo-font--proPlace']/../../div[@class='woo-box-item-flex ProfileHeader_con3_Bg19p']/text()").extract_first()

    # 获取用户IP
    ip              = response.xpath("//div[@class='woo-panel-main woo-panel-top woo-panel-right woo-panel-bottom woo-panel-left Card_wrap_2ibWe Card_bottomGap_2Xjqi']//i[@class='woo-font woo-font--ip']/../..//div[@class='woo-box-item-flex ProfileHeader_con3_Bg19p' and @style]/text()").extract_first()

    # 获取加入微博时间
    join_date       = response.xpath("//div[@class='woo-panel-main woo-panel-top woo-panel-right woo-panel-bottom woo-panel-left Card_wrap_2ibWe Card_bottomGap_2Xjqi']//i[@class='woo-font woo-font--proTime']/../../div[@class='woo-box-item-flex ProfileHeader_con3_Bg19p']/text()").extract_first()

    # 获取用户信用
    credit          = response.xpath("//div[@class='woo-panel-main woo-panel-top woo-panel-right woo-panel-bottom woo-panel-left Card_wrap_2ibWe Card_bottomGap_2Xjqi']//i[@class='woo-font woo-font--proCredit']/../../div[@class='woo-box-item-flex ProfileHeader_con3_Bg19p']/text()").extract_first()

    # 获取出生日期、星座
    birth           = response.xpath("//div[@class='woo-panel-main woo-panel-top woo-panel-right woo-panel-bottom woo-panel-left Card_wrap_2ibWe Card_bottomGap_2Xjqi']//i[@class='woo-font woo-font--proIntro']/../../div[@class='woo-box-item-flex ProfileHeader_con3_Bg19p']/span/text()").extract_first()

    # 获取v认证
    v_auth          = response.xpath("//div[@class='woo-panel-main woo-panel-top woo-panel-right woo-panel-bottom woo-panel-left Card_wrap_2ibWe Card_bottomGap_2Xjqi']//i[@class='woo-font woo-font--proV']/../../div[@class='woo-box-item-flex ProfileHeader_con3_Bg19p ProfileHeader_flexBasisAuto_2exBQ ProfileHeader_descText_3AF6o']/text()").extract_first()

    # 获取身份认证
    id_auth         = response.xpath("//div[@class='woo-panel-main woo-panel-top woo-panel-right woo-panel-bottom woo-panel-left Card_wrap_2ibWe Card_bottomGap_2Xjqi']//span[@class='woo-icon-wrap woo-avatar-icon']/@title").extract_first()

    # 获取表现
    performance     = '    '.join(response.xpath("//div[@class='woo-panel-main woo-panel-top woo-panel-right woo-panel-bottom woo-panel-left Card_wrap_2ibWe Card_bottomGap_2Xjqi']//div[@class='woo-box-flex woo-box-alignCenter woo-box-justifyBetween']/div/div/*/text()").extract())

    # 获取内容/IP机构认证
    content_auth    = response.xpath("//div[@class='woo-panel-main woo-panel-top woo-panel-right woo-panel-bottom woo-panel-left Card_wrap_2ibWe Card_bottomGap_2Xjqi']//i[@class='woo-font woo-font--proCom']/../../div[@class='woo-box-item-flex ProfileHeader_con3_Bg19p']/text()").extract_first()

    # 获取校园
    campus          = response.xpath("//div[@class='woo-panel-main woo-panel-top woo-panel-right woo-panel-bottom woo-panel-left Card_wrap_2ibWe Card_bottomGap_2Xjqi']//i[@class='woo-font woo-font--proEdu']/../../div[@class='woo-box-item-flex ProfileHeader_con3_Bg19p']/text()").extract_first()

    # 获取合作机构
    cooperative_org = response.xpath("//div[@class='woo-panel-main woo-panel-top woo-panel-right woo-panel-bottom woo-panel-left Card_wrap_2ibWe Card_bottomGap_2Xjqi']//i[@class='woo-font woo-font--mcn']/../../div[@class='woo-box-item-flex ProfileHeader_con3_Bg19p']/text()").extract_first()
    
    # 保存头像
    with open(f"{graphs_dir}/{uid}/avatar.jpg", "wb") as file: 
        try:
            WebDriverWait(driver, 0.15, 0.001).until(EC.presence_of_element_located((By.XPATH, "//div[@class='woo-avatar-main woo-avatar-hover ProfileHeader_avatar2_1gEyo']/img")))
            avatar = driver.find_element(By.XPATH, "//div[@class='woo-avatar-main woo-avatar-hover ProfileHeader_avatar2_1gEyo']/img")
            avatar_url = avatar.get_attribute('src')
            content = requests.get(avatar_url).content
            file.write(content);
        except:
            traceback
    avatar_url      = f"{graphs_dir}/{uid}/avatar.jpg"

    # 保存背景
    with open(f"{graphs_dir}/{uid}/background.jpg", "wb") as file: 
        try:
            WebDriverWait(driver, 0.15, 0.001).until(EC.presence_of_element_located((By.XPATH, "//div[@class='woo-picture-main ProfileHeader_pic_2Coeq']/img")))
            background = driver.find_element(By.XPATH, "//div[@class='woo-picture-main ProfileHeader_pic_2Coeq']/img")
            background_url = background.get_attribute('src')
            content = requests.get(background_url).content
            file.write(content)
        except:
            traceback
    background_url  = f"{graphs_dir}/{uid}/background.jpg"

    # 返回获取到用户的全部数据
    return [avatar_url,     background_url,     username,       gender,         
            level,          followers,          followings,     weibo_count,    
            introduction,   address,            ip,             join_date,     
            credit,         birth,              v_auth,         id_auth,       
            performance,    content_auth,       campus,         cooperative_org]

# %% [markdown]
# ```python
# 'get_weibo(driver, num)'
# ```
# 参数|说明
# --|--
# driver|浏览器驱动
# num|要爬取的微博数量
# 说明 | 只适用于微博网页<a>https://weibo.com</a>

# %%
def get_weibos(driver, href, num):

    weibo_dic   =   list()

    while driver.current_url != href:
        try: driver.get(href); break
        except: traceback
        
        try: driver.get(random.choice(['https://weibo.com', 'https://weibo.com/at/weibo']))
        except: traceback
        sleep(random.randint(20, 60))


    for prog in tqdm(range(num), bar_format=href+' 微博数据爬取中:\t{l_bar}{bar:30}{r_bar}'):

        """找到微博容器的列表"""
        while True:
            try:
                WebDriverWait(driver, 0.3, 0.001).until(EC.presence_of_element_located((By.XPATH, "//div[@class='vue-recycle-scroller__item-view']")))
                elements = [e for e in driver.find_elements(By.XPATH, "//div[@class='vue-recycle-scroller__item-view']") if e != None]
                if elements: break
            except: traceback

        """
            遍历每一个容器、展开图片会改变当前位置
            但是展开文本并不会、本次遍历目的是展开文本
        """
        for elm in elements:
            try:
                WebDriverWait(driver, 0.15, 0.001).until(EC.presence_of_element_located((By.XPATH, ".//div[@class='detail_wbtext_4CRf9']//span")))
                more = [e for e in elm.find_elements(By.XPATH, ".//div[@class='detail_wbtext_4CRf9']//span") if e.text == '展开' and e != None]
                for e in more:
                    try:    e.click();
                    except: traceback
            except: traceback

        response = Selector(text=driver.page_source)

        """遍历每一个容器"""
        for elm in response.xpath("//div[@class='vue-recycle-scroller__item-view']"):

            """声明变量"""
            uid,        date,           dev,            content,        img,                \
            fwd_count,  cmt_count,      lk_count,       is_fwd,         is_fast_fwd,        \
            is_like,    src_uid,        src_uname,      src_date,       src_dev,            \
            src_cont,   src_img,        src_fwd_count,  src_cmt_count,  src_lk_count    =   [None]*20
            
            uid         = href.replace('https://weibo.com/u/', '')
            is_fwd      = False
            is_fast_fwd = False
            is_like     = False
            add_img_src = None
            img_srcs    = set()

            src_wb      = elm.xpath(".//div[@class='Feed_retweet_JqZJb']")
            """********************************************************************************************************************************"""            
            """判断是否为转发微博"""
            if  src_wb:
                is_fwd      = True

            """判断是否为快转的微博"""
            if elm.xpath(".//span[@class='head_fastbehind_1StRl']").extract(): 
                is_fast_fwd = True

            """判断是否为点赞的微博"""
            if elm.xpath(".//span[@class='title_title_1DVuO']").extract() and not elm.xpath(".//span[@class='title_title_1DVuO']/img").extract():
                is_like     = True
                is_fwd      = False

            """图片"""
            img_srcs = img_srcs | set(elm.xpath(".//div[@class='picture picture-box_row_30Iwo']//img/@src").extract())

            """"创建目录"""
            shutil.rmtree(f"{graphs_dir}/{uid}/{len(weibo_dic)}")           if     os.path.exists(f"{graphs_dir}/{uid}/{len(weibo_dic)}") else None
            os.makedirs  (f"{graphs_dir}/{uid}/{len(weibo_dic)}", mode=777) if not os.path.exists(f"{graphs_dir}/{uid}/{len(weibo_dic)}") else None

            """根据不同类型的微博给变量赋值****************************************************************************************************************"""
            if is_fwd:
                src_uid     = src_wb.xpath(".//a[@class='ALink_default_2ibt1']/@href").extract_first()

                if not src_uid:
                    src_uid = src_wb.xpath(".//a[@class='router-link-exact-active router-link-active ALink_default_2ibt1']/@href").extract_first()
                    
                if src_uid:
                    src_uid = src_uid.replace('https://weibo.com/u/', '').replace('/u/', '')

                src_uname   = src_wb.xpath(".//span[@class='detail_nick_u-ffy']/text()").extract_first()

                if src_uname:
                    src_uname = src_uname.replace('@', '')
                
                src_date    = src_wb.xpath(".//a[@class='head-info_time_6sFQg']/@title").extract_first()

                src_cont    = '    '.join(src_wb.xpath(".//div[@class='detail_wbtext_4CRf9']/text()").extract()).replace('\n', '    ')

                src_img     = f"{graphs_dir}/{uid}/{len(weibo_dic)}"

                add_img_src = elm.xpath(".//div[@class='detail_wbtext_4CRf9']/a[@target='_blank']/@href").extract_first()

                img         = f"{graphs_dir}/{uid}/{len(weibo_dic)}_addition"

                date        = elm.xpath(".//header/div/div/div/a[@class='head-info_time_6sFQg']/@title").extract_first()

                dev         = elm.xpath(".//div[@class='head-info_cut_1tPQI head-info_source_2zcEX']/text()").extract_first()

                other_dev   = [text for text in elm.xpath(".//div[@class='head-info_cut_1tPQI head-info_source_2zcEX']//*/text()").extract() if text != None and text != '']
                
                if other_dev:
                    dev     = dev + ' ' + ' '.join(other_dev)

                content     = '    '.join(elm.xpath(".//article//div[@class='detail_text_1U10O detail_ogText_2Z1Q8 wbpro-feed-ogText']//div[@class='detail_wbtext_4CRf9']/text()").extract()).replace('\n', '    ')

                line        = src_wb.xpath(".//footer/@aria-label").extract_first()
                
                if line:
                    src_fwd_count, src_cmt_count, src_lk_count = line.split(',')

                line        = elm.xpath(".//article/footer/@aria-label").extract_first()

                if line:
                    fwd_count,     cmt_count,     lk_count     = line.split(',')
                 
            # ***************************************************************************************************************************
            elif is_like or is_fast_fwd:
                src_uid     = elm.xpath(".//a[@class='ALink_default_2ibt1']/@href").extract_first()

                if src_uid:
                    src_uid = src_uid.replace('https://weibo.com/u/', '').replace('/u/', '')

                src_uname   = elm.xpath(".//a[@class='ALink_default_2ibt1']/@aria-label").extract_first()

                if src_uname:
                    src_uname = src_uname.replace('@', '')

                src_date    = elm.xpath(".//header/div/div/div/a[@class='head-info_time_6sFQg']/@title").extract_first()

                if is_like:
                    date    = elm.xpath(".//span[@class='title_title_1DVuO']/text()").extract_first()

                src_dev     = elm.xpath(".//div[@class='head-info_cut_1tPQI head-info_source_2zcEX']/text()").extract_first()

                other_dev   = [text for text in elm.xpath(".//div[@class='head-info_cut_1tPQI head-info_source_2zcEX']//*/text()").extract() if text != None and text != '']
                
                if other_dev:
                    src_dev = src_dev + ' ' + ' '.join(other_dev)
                
                src_cont    = '    '.join(elm.xpath(".//article//div[@class='detail_text_1U10O detail_ogText_2Z1Q8 wbpro-feed-ogText']//div[@class='detail_wbtext_4CRf9']/text()").extract()).replace('\n', '    ')

                src_img     = f"{graphs_dir}/{uid}/{len(weibo_dic)}"

                line        = elm.xpath(".//article/footer/@aria-label").extract_first()

                if is_like and line:
                    src_fwd_count, src_cmt_count, src_lk_count = line.split(',')

                elif line:
                    fwd_count,     cmt_count,     lk_count     = line.split(",")
                
            # ***********************************************************************************************************************
            else:     
                date        = elm.xpath(".//header/div/div/div/a[@class='head-info_time_6sFQg']/@title").extract_first()

                dev         = elm.xpath(".//div[@class='head-info_cut_1tPQI head-info_source_2zcEX']/text()").extract_first()

                other_dev   = [text for text in elm.xpath(".//div[@class='head-info_cut_1tPQI head-info_source_2zcEX']//*/text()").extract() if text != None and text != '']
                
                if other_dev:
                    dev     = dev + ' ' + ' '.join(other_dev)

                content     = '    '.join(elm.xpath(".//article//div[@class='detail_text_1U10O detail_ogText_2Z1Q8 wbpro-feed-ogText']//div[@class='detail_wbtext_4CRf9']/text()").extract()).replace('\n', '    ')

                img = f"{graphs_dir}/{uid}/{len(weibo_dic)}"
                
                line        = elm.xpath(".//article/footer/@aria-label").extract_first()

                if line:
                    fwd_count, cmt_count, lk_count = line.split(',')

            """**************************************************************************************************************************"""
            """添加数据"""
            data_dic = {'uid'           :uid,
                        'date'          :date,
                        'dev'           :dev,
                        'content'       :content,
                        'img'           :img,
                        'fwd_count'     :fwd_count,
                        'cmt_count'     :cmt_count,
                        'lk_count'      :lk_count,
                        'is_fwd'        :is_fwd,
                        'is_fast_fwd'   :is_fast_fwd,
                        'is_like'       :is_like,
                        'src_uid'       :src_uid,
                        'src_uname'     :src_uname,
                        'src_date'      :src_date,
                        'src_dev'       :src_dev,
                        'src_cont'      :src_cont,
                        'src_img'       :src_img,
                        'src_fwd_count' :src_fwd_count,
                        'src_cmt_count' :src_cmt_count,
                        'src_lk_count'  :src_lk_count}

            can_add = True if '' not in list(data_dic.values()) else False

            if can_add:
                for dic in copy.deepcopy(weibo_dic):
                    dt_dic_cp = copy.deepcopy(data_dic)
                    dt_dic_cp.pop('img')
                    dt_dic_cp.pop('src_img')
                    dic.pop('img')
                    dic.pop('src_img')
                    can_add = True if list(dic.values()) != list(dt_dic_cp.values()) else False
                    if not can_add: break
                    
            if  can_add:
                for index, src in enumerate(img_srcs):
                    with open(f"{graphs_dir}/{uid}/{len(weibo_dic)}/{index}.jpg", 'wb') as file:
                        try: file.write(requests.get(src).content)
                        except: traceback
                
                if is_fwd and img and add_img_src:
                    shutil.rmtree(img)  if     os.path.exists(img)  else None
                    os.makedirs(img)    if not os.path.exists(img)  else None
                    with open(f"{img}/addition.jpg", 'wb') as file:
                        try: file.write(requests.get(add_img_src).content)
                        except: traceback

                weibo_dic.append(data_dic)
            
        """向下滚动900px"""
        driver.execute_script("window.scrollBy(0, 1080);")

    return [list(dic.values()) for dic in weibo_dic]

# %% [markdown]
# ```python
# 'crawling(driver, lanbel, num)'
# ```
# 参数|说明
# ---|---
# driver | 浏览器驱动
# label | 要爬取的标签
# num | 要在当前专题下爬取的人数
# 说明 | 只适用于微博网页<a>https://weibo.com</a>

# %%
def crawling(driver, label, num=60):

    usr_href_set = set()

    try:
        WebDriverWait(driver, 1.5, 0.001).until(EC.presence_of_element_located((By.XPATH, "//div[@class='woo-box-flex woo-box-alignCenter woo-box-justifyCenter Ctrls_item_3KzNH' and @title='热门']")))
        driver.find_element(By.XPATH, "//div[@class='woo-box-flex woo-box-alignCenter woo-box-justifyCenter Ctrls_item_3KzNH' and @title='热门']").click()
    except:
        traceback

    try:
        WebDriverWait(driver, 1.5, 0.001).until(EC.presence_of_element_located((By.XPATH, "//i[@class='woo-font woo-font--more' and @title='更多']")))
        driver.find_element(By.XPATH, "//i[@class='woo-font woo-font--more' and @title='更多']").click()
    except:
        traceback

    # 找到多个专题、然后点击名称为label的专题
    while True:
        sleep(0.15)
        try:
            e = [e for e in driver.find_elements(By.XPATH,  "//div[@class='wbpro-textcut']") if e.text==label and e!=None].pop()
            e.click(); break
        except: 
            traceback

    # 持续向下滚动、扫描到num个用户的主页href则停止
    while len(usr_href_set) < num:
        sleep(0.15)
        try:
            usr_href_set = usr_href_set | {item.get_attribute('href') for item in driver.find_elements(By.XPATH, "//a[@class='ALink_default_2ibt1 head_cut_2Zcft head_name_24eEB']") if item.text!=''}
            driver.execute_script("window.scrollBy(0, 360);")
        except:
            traceback
        while len(usr_href_set) > num: usr_href_set.pop()

    """"******************************************************************************************************************************************"""
    """遍历每一位user_href_set里的用户"""
    for href in usr_href_set:
        uid             = href.replace('https://weibo.com/u/', '')
        follow_href_set = set()
        follow          = None
        num             = None

        # 先进入用户主页
        while driver.current_url != href:
            try: driver.get(href); break
            except: traceback
            
            try: driver.get(random.choice(['https://weibo.com', 'https://weibo.com/at/weibo']))
            except: traceback
            sleep(random.randint(20, 60))
        
        # 关注此用户
        while True:
            try:
                WebDriverWait(driver, 1.5, 0.001).until(EC.presence_of_element_located((By.XPATH, "//button[@class='woo-button-main woo-button-flat woo-button-primary woo-button-m woo-button-round FollowBtn_m_1UJhp ProfileHeader_btn3_2VD_Y']")))
                driver.find_element(By.XPATH, "//button[@class='woo-button-main woo-button-flat woo-button-primary woo-button-m woo-button-round FollowBtn_m_1UJhp ProfileHeader_btn3_2VD_Y']").click()
                sleep(0.6)
                break
            except:
                traceback
        
        # 获取关注数
        while follow in {None, ''}:
            follow = Selector(text=driver.page_source).xpath("//div[@class='woo-box-flex woo-box-alignCenter ProfileHeader_h4_gcwJi']/a[@class='ALink_none_1w6rm ProfileHeader_alink_tjHJR ProfileHeader_pointer_2yKGQ']/span/span/text()").extract_first()

        if  follow.isdigit():
            follow = 250 if int(follow) > 250 else int(follow)
        else:
            follow = 250

        # 获取微博数量
        while num in {None, ''}:
            num = Selector(text=driver.page_source).xpath("//div[@class='wbpro-screen-v2 woo-box-flex woo-box-alignCenter woo-box-justifyBetween']/div/text()").extract_first()
        
        num     = int(''.join([c for c in num if c.isdigit()]))
        num     = 1000 if num > 1000 else num

        # 进入用户的关注主页
        while True:
            try:
                driver.find_element(By.XPATH, "//div[@class='woo-box-flex woo-box-alignCenter ProfileHeader_h4_gcwJi']/a[@class='ALink_none_1w6rm ProfileHeader_alink_tjHJR ProfileHeader_pointer_2yKGQ']").click()
                break
            except:
                traceback
            
        sleep(0.6)

        # 获取到当前用户的所有关注的主页面链接
        while follow > 0:
            try:
                follow_href_set = follow_href_set | {item.get_attribute('href') for item in driver.find_elements(By.XPATH, "//a[@class='ALink_none_1w6rm UserCard_item_TrVS0']")}
                driver.execute_script("window.scrollBy(0, 188);")
            except:
                traceback
            follow -= 2

        # 新建用户保存图片的目录
        for item in {href} | follow_href_set:
            dir_path = graphs_dir + '/' + item.replace('https://weibo.com/u/', '')
            os.makedirs(dir_path) if not os.path.exists(dir_path) else None

        # 添加用户自身的数据
        add_to_excel(user_profiles, [[uid, label] + get_user_fields(driver, href)])

        # 添加用户的微博数据
        add_to_excel(weibo_profiles, get_weibos(driver, href, num))

        driver.execute_script("window.scrollTo(0, 0);")
        sleep(1.5)

        # 鼠标悬停到已关注button上
        try:
            WebDriverWait(driver, 3.0, 0.001).until(EC.presence_of_element_located((By.XPATH, "//button[@class='woo-button-main woo-button-line woo-button-default woo-button-m woo-button-round FollowBtn_m_1UJhp ProfileHeader_btn3_2VD_Y']")))
            element = driver.find_element(By.XPATH, "//button[@class='woo-button-main woo-button-line woo-button-default woo-button-m woo-button-round FollowBtn_m_1UJhp ProfileHeader_btn3_2VD_Y']")
            ActionChains(driver).move_to_element(element).perform()
        except:
            traceback

        # 点击取消关注
        try:
            WebDriverWait(driver, 3.0, 0.001).until(EC.presence_of_element_located((By.XPATH, "//div[@class='woo-box-flex woo-box-alignCenter woo-pop-item-main FollowPop_item_1GgQ0' and @role='button']")))
            elements = driver.find_elements(By.XPATH, "//div[@class='woo-box-flex woo-box-alignCenter woo-pop-item-main FollowPop_item_1GgQ0' and @role='button']")
            [e.click() for e in elements if '取消关注' in e.text]
        except:
            traceback

        # 确认
        try:
            WebDriverWait(driver, 3.0, 0.001).until(EC.presence_of_element_located((By.XPATH, "//button[@class='woo-button-main woo-button-flat woo-button-primary woo-button-m woo-button-round woo-dialog-btn']")))
            element = driver.find_element(By.XPATH, "//button[@class='woo-button-main woo-button-flat woo-button-primary woo-button-m woo-button-round woo-dialog-btn']")
            element.click()
        except:
            traceback

        while True:
            sleep(0.6)
            try: driver.find_element(By.XPATH, "//div[@class='woo-box-flex woo-box-column woo-box-alignCenter woo-box-justifyCenter woo-toast-main woo-toast--success']/div[@class='woo-toast-body']/span[text()='取关成功']")
            except: traceback; break
        
        # 遍历获取到的链接、拿到此用户关注的每一位用户的基本数据
        temp_data_lst = list()
        for prog in tqdm(range(len(follow_href_set)), bar_format=href+' 用户关注爬取中:\t{l_bar}{bar:30}{r_bar}'):
            item = follow_href_set.pop()
            temp_data_lst.append([uid, item.replace('https://weibo.com/u/', '')] + get_user_fields(driver, item))

        add_to_excel(follow_profiles, temp_data_lst)

# %% [markdown]
# - 执行

# %%
if __name__ == '__main__':

    # 添加xlsx文件属性
    for item in [(user_profiles,user_fields), (follow_profiles,follow_fields), (weibo_profiles,weibo_fields)]:
        path, cont = item
        book       = openpyxl.load_workbook(path)
        sheet      = book.active

        if sheet.max_row == 1 and sheet.max_column == 1: 
            add_to_excel(path, [cont])

    # 设置标签
    labels = ['艺术', '明星', '动漫', '搞笑', '游戏', 
              '美妆', '美食', '摄影', '数码', '萌宠', 
              '星座', '体育', '旅游', '教育', '国际', 
              '财经', '婚恋', '时尚', '科技', '校园']

    # 更新cookies
    with open(cookies_txt, 'r+') as file:
        if len(file.readlines()) == 0:  
            update_cookies(cookies_txt)

    # 创建 FirefoxOptions 对象，设置无头模式和允许加载图片
    options             = webdriver.FirefoxOptions()
    options.headless    = True
    options.set_preference("permissions.default.image", 0)

    driver = webdriver.Firefox(options=options)

    sleep(5)

    # 进入登录界面
    driver.get('https://weibo.com/login.php')

    # 等待登录成功
    while driver.current_url.find('https://weibo.com/login.php') == -1: sleep(3)

    # 设置cookies
    while True:
        sleep(3);       set_cookies(driver, cookies_txt)
        sleep(1);       driver.refresh()    
        sleep(3);

        if  driver.current_url.find('https://weibo.com/login.php') != -1:
            update_cookies(cookies_txt)

        else: break


    # 爬虫主程序
    while True:
        operation = input('\n是否爬虫(yes or no):\t')

        if operation in {'y', 'Y', 'yes', 'YES'}:
            driver.set_page_load_timeout(10)
            titles = list()
            num    = 0

            for index, topic in enumerate(labels, start=1):
                print(topic, end='\n', flush=True) if index % 5 == 0 else print(topic, end='\t', flush=True)
                
            while True:
                for title in input('输入要爬取的专题(以空格分隔):\t').split(' '):
                    titles.append(title) if title in labels else print(f'{title}不在专题范围内', flush=True)

                if len(titles) > 0: break
                
                print('没有匹配项', flush=True)


            while True:
                num = input('输入要爬取的人数(请输入大于0的整数):\t')
                if not num.isdigit(): print('请输入大于0的整数', flush=True)
                elif int(num) <= 0: print('请输入大于0的整数', flush=True)
                else: num = int(num); break
            
            for title in titles:
                crawling(driver, title, num)

            print('爬取完毕', flush=True)

        elif operation in {'n', 'N', 'no', 'NO'}: break
        else: print('指令错误', flush=True)

    # 关闭爬虫程序
    driver.close()
    driver.quit()

# %% [markdown]
# - 打包程序
# - pip install pyinstaller
# - pyinstaller -F --hidden-import=openpyxl.cell._writer  file_name

# %% [markdown]
# #### `Test`

# %%
# # 添加xlsx文件属性
# for item in [(user_profiles,user_fields), (follow_profiles,follow_fields), (weibo_profiles,weibo_fields)]:
#     path, cont = item
#     book       = openpyxl.load_workbook(path)
#     sheet      = book.active

#     if sheet.max_row == 1 and sheet.max_column == 1: 
#         add_to_excel(path, [cont])

# # 设置标签
# labels = ['艺术', '明星', '动漫', '搞笑', '游戏', 
#           '美妆', '美食', '摄影', '数码', '萌宠', 
#           '星座', '体育', '旅游', '教育', '国际', 
#           '财经', '婚恋', '时尚', '科技', '校园']

# # 更新cookies
# with open(cookies_txt, 'r+') as file:
#     if len(file.readlines()) == 0:  
#         update_cookies(cookies_txt)

# driver = webdriver.Firefox()

# # 进入登录界面
# driver.get('https://weibo.com/login.php')

# # 等待登录成功
# while driver.current_url.find('https://weibo.com/login.php') == -1: sleep(3)

# # 设置cookies
# while True:
#     sleep(3);       set_cookies(driver, cookies_txt)
#     sleep(1);       driver.refresh()    
#     sleep(3);

#     if  driver.current_url.find('https://weibo.com/login.php') != -1:
#         update_cookies(cookies_txt)

#     else: break

# crawling(driver, '美妆', 2)


